from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

import openpyxl


TARGET_SHEET = "DataBase (IMO form)"


def _clean(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def load_arrival_database(excel_path: str) -> List[Dict[str, str]]:
    path = Path(excel_path)
    wb = openpyxl.load_workbook(path, data_only=True)

    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{TARGET_SHEET}' not found in {excel_path}")

    ws = wb[TARGET_SHEET]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_clean(x) for x in rows[0]]
    crew_rows: List[Dict[str, str]] = []

    for row in rows[1:]:
        record = {}
        for i, value in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = _clean(value)

        if any(record.values()):
            crew_rows.append(record)

    return crew_rows
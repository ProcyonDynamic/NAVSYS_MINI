import os
from openpyxl import load_workbook


EXCEL_PATH = r"D:\NAVSYS_USB\data\PORTALIS\arrival_workbook.xlsx"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value if value else None
    return value


def sheet_map_by_headers(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    items = []

    for row in rows[1:]:
        if row is None:
            continue

        item = {}
        has_data = False

        for i, header in enumerate(headers):
            if not header:
                continue

            value = row[i] if i < len(row) else None
            value = clean(value)

            if value is not None:
                has_data = True

            item[header] = value

        if has_data:
            items.append(item)

    return items


def first_sheet_row_dict(ws):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    values = rows[1]

    item = {}
    for i, header in enumerate(headers):
        if not header:
            continue
        value = values[i] if i < len(values) else None
        item[header] = clean(value)

    return item


def pick_first_nonempty(row_dict, *keys):
    for key in keys:
        value = row_dict.get(key)
        if value not in (None, ""):
            return value
    return None


def build_context_from_workbook(excel_path=EXCEL_PATH, selected_crew_index=0):
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel workbook not found: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)

    context = {
        "crew": {},
        "ship": {},
        "voyage": {},
        "cargo": {},
        "company": {}
    }

    # Expected sheet names you can change later
    ship_ws = wb["SHIP"] if "SHIP" in wb.sheetnames else None
    voyage_ws = wb["VOYAGE"] if "VOYAGE" in wb.sheetnames else None
    cargo_ws = wb["CARGO"] if "CARGO" in wb.sheetnames else None
    crew_ws = wb["CREW"] if "CREW" in wb.sheetnames else None
    company_ws = wb["COMPANY"] if "COMPANY" in wb.sheetnames else None

    if ship_ws:
        row = first_sheet_row_dict(ship_ws)
        context["ship"] = {
            "name": pick_first_nonempty(row, "Ship Name", "Vessel Name", "Name"),
            "imo": pick_first_nonempty(row, "IMO", "IMO Number"),
            "flag": pick_first_nonempty(row, "Flag", "Flag State"),
            "call_sign": pick_first_nonempty(row, "Call Sign", "Callsign"),
            "mmsi": pick_first_nonempty(row, "MMSI", "MMSI Number"),
            "port_of_registry": pick_first_nonempty(row, "Port of Registry", "Registry Port"),
            "owner": pick_first_nonempty(row, "Owner", "Registered Owner"),
            "operator": pick_first_nonempty(row, "Operator", "Company", "Manager"),
            "master": pick_first_nonempty(row, "Master", "Master Name")
        }

    if voyage_ws:
        row = first_sheet_row_dict(voyage_ws)
        context["voyage"] = {
            "last_port": pick_first_nonempty(row, "Last Port", "Previous Port", "Port of Departure"),
            "next_port": pick_first_nonempty(row, "Next Port", "Arrival Port", "Port of Arrival"),
            "eta": pick_first_nonempty(row, "ETA", "Estimated Time of Arrival"),
            "etd": pick_first_nonempty(row, "ETD", "Estimated Time of Departure"),
            "agent": pick_first_nonempty(row, "Agent", "Port Agent", "Shipping Agent"),
            "reporting_party": pick_first_nonempty(row, "Reporting Party", "Submitting Party"),
            "notice_id": pick_first_nonempty(row, "Notice ID", "Submission ID"),
            "berth": pick_first_nonempty(row, "Berth", "Berth No"),
            "terminal": pick_first_nonempty(row, "Terminal", "Terminal Name")
        }

    if cargo_ws:
        row = first_sheet_row_dict(cargo_ws)
        context["cargo"] = {
            "type": pick_first_nonempty(row, "Cargo Type", "Cargo", "Commodity"),
            "quantity": pick_first_nonempty(row, "Cargo Quantity", "Quantity", "Cargo Amount"),
            "quantity_on_board": pick_first_nonempty(row, "Quantity on Board", "QOB"),
            "temperature": pick_first_nonempty(row, "Cargo Temperature", "Cargo Temp", "Tank Temperature"),
            "boil_off_rate": pick_first_nonempty(row, "BOR", "Boil Off Rate", "Daily Boil Off")
        }

    if company_ws:
        row = first_sheet_row_dict(company_ws)
        context["company"] = {
            "dpa": pick_first_nonempty(row, "DPA", "Designated Person Ashore"),
            "cso": pick_first_nonempty(row, "CSO", "Company Security Officer")
        }

    if crew_ws:
        crew_rows = sheet_map_by_headers(crew_ws)

        if crew_rows:
            selected_crew_index = max(0, min(selected_crew_index, len(crew_rows) - 1))
            crew_row = crew_rows[selected_crew_index]

            context["crew"] = {
                "first_name": pick_first_nonempty(crew_row, "First Name", "Given Name"),
                "last_name": pick_first_nonempty(crew_row, "Last Name", "Surname", "Family Name"),
                "middle_name": pick_first_nonempty(crew_row, "Middle Name", "Other Names"),
                "nationality": pick_first_nonempty(crew_row, "Nationality", "Citizenship"),
                "date_of_birth": pick_first_nonempty(crew_row, "Date of Birth", "DOB"),
                "place_of_birth": pick_first_nonempty(crew_row, "Place of Birth", "Birth Place"),
                "gender": pick_first_nonempty(crew_row, "Gender", "Sex"),
                "rank": pick_first_nonempty(crew_row, "Rank", "Position"),
                "home_address": pick_first_nonempty(crew_row, "Home Address", "Address", "Permanent Address"),
                "passport": {
                    "number": pick_first_nonempty(crew_row, "Passport Number", "Passport No", "Travel Document Number"),
                    "expiry_date": pick_first_nonempty(crew_row, "Passport Expiry", "Passport Expiry Date"),
                    "issue_date": pick_first_nonempty(crew_row, "Passport Issue Date", "Date of Issue"),
                    "place_of_issue": pick_first_nonempty(crew_row, "Passport Place of Issue", "Place of Issue"),
                    "country_of_issue": pick_first_nonempty(crew_row, "Passport Country of Issue", "Country of Issue")
                },
                "visa": {
                    "number": pick_first_nonempty(crew_row, "Visa Number", "Visa No"),
                    "expiry_date": pick_first_nonempty(crew_row, "Visa Expiry", "Visa Expiry Date"),
                    "type": pick_first_nonempty(crew_row, "Visa Type", "Visa Category")
                },
                "seaman_book": {
                    "number": pick_first_nonempty(crew_row, "Seaman Book Number", "CDC Number", "Seafarer Book"),
                    "expiry_date": pick_first_nonempty(crew_row, "Seaman Book Expiry", "CDC Expiry")
                }
            }

    return context